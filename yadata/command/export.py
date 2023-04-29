import sys
import yadata.utils.sane_yaml as sane_yaml
from yadata.command.command import YadataCommand
from yadata.utils.misc import Argument

from collections import defaultdict


def is_atomic(x):

    return type(x) in (str,int,float)

def addr(row,col):

    return chr(ord('A')+col)+str(row+1)

class Export(YadataCommand):
    """reads stream of records, outputs Excel spreadsheet
"""

    name="render"

    arguments=(
        Argument("outfile",help="Excel filename for output"),
    )

    data_in=True
    data_out=False

    def __init__(self,ns):
        self.ns=ns

    def execute(self,it):

        # An expensive import belongs here
        from openpyxl import Workbook
        records=list(it)
        workbook=Workbook()
        del workbook[workbook.active.title]
        tags_dict=defaultdict(lambda:[])
        for rec in records:
           tags_dict[rec.yadata_tag].append(rec)
        for tag,records_with_tag in tags_dict.items():
            worksheet=workbook.create_sheet(title=tag)
            keys=[]
            inverse_keys={}
            for rec in records_with_tag:
                for key,value in rec.items():
                    if not is_atomic(value):
                        if key in keys:
                            keys.remove(key)
                        continue
                    if key not in keys:
                        keys.append(key)
            for i,key in enumerate(keys):
                inverse_keys[key]=i
                #worksheet.write(0,i,key)
                worksheet.cell(row=1,column=i+1,value=key)
            for row,rec in enumerate(records_with_tag,start=1):
                for key,value in rec.items():
                    if key not in inverse_keys:
                        continue
                    try:
                        worksheet.cell(row+1,column=inverse_keys[key]+1,value=value)
                    except:
                        print(key,value)
                        raise
        workbook.save(self.ns.outfile)


                    

            

